from __future__ import annotations

import re
import shutil
import tempfile
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


STATEMENTS = ("Income Statement", "Balance Sheet", "Cash Flow Statement")

INCOME_KEYS = (
    "revenue",
    "sale",
    "sales",
    "income",
    "expense",
    "expenses",
    "cost",
    "profit",
    "loss",
    "tax",
    "earnings",
    "share",
    "comprehensive",
    "depreciation",
    "amortisation",
    "amortization",
    "finance",
)

BALANCE_KEYS = (
    "asset",
    "assets",
    "property",
    "equipment",
    "capital work",
    "investment",
    "goodwill",
    "intangible",
    "inventory",
    "inventories",
    "receivable",
    "cash",
    "bank",
    "equity",
    "capital",
    "liabilit",
    "borrowings",
    "lease",
    "payable",
    "provision",
    "deferred tax",
)

CASH_FLOW_KEYS = (
    "cash",
    "operating activities",
    "investing activities",
    "financing activities",
    "depreciation",
    "amortization",
    "amortisation",
    "stock-based",
    "inventories",
    "receivable",
    "payable",
    "purchase",
    "proceeds",
    "repayment",
    "borrowings",
    "dividend",
    "interest",
    "tax paid",
    "net cash generated",
    "net cash provided",
    "exchange",
    "lease",
)

SUMMARY_LABELS = {
    "Revenue": (
        "revenue from operations",
        "total net sales",
        "total revenue from operations",
        "net sales",
    ),
    # Covers profit- and loss-making wording, US and Indian.
    "PBT": (
        "profit before tax", "loss before tax", "profit/(loss) before tax",
        "income before income taxes", "before income taxes", "before tax",
    ),
    "PAT": (
        "profit for the year", "loss for the year", "profit for the period",
        "loss for the period", "profit/(loss) for the", "net income",
        "net profit for the", "net profit/(loss)",
    ),
    "Total Assets": ("total assets",),
    "Total Equity": ("total equity", "total stockholders' equity", "total shareholders' equity"),
    # Many Indian reports word the net operating line differently ("Net cash
    # inflow from … operating activities", "Net cash flows generated from …").
    "Operating Cash Flow": (
        "net cash generated from operating activities",
        "net cash provided by operating activities",
        "net cash provided by (used in) operating activities",
        "net cash flows generated from operating activities",
        "net cash flow from operating activities",
        "net cash flows from operating activities",
        "net cash from operating activities",
        "net cash inflow from",
        "net cash used in operating activities",
        "net cash generated from",  # "...from / (Used In) Operating Activities (A)"
        "operating activities (a)",  # Indian net-operating lines suffix the section "(A)"
    ),
}


@dataclass
class ParsedPdf:
    company: str
    source_name: str
    source_path: Path
    periods: list[str] = field(default_factory=list)
    # statements[stmt][canonical_key] = {period: value}. The canonical key is used
    # only to MERGE the same line across years; the verbatim label shown in the
    # sheet is kept in display_labels, and the PDF row position in row_order.
    statements: dict[str, dict[str, dict[str, float]]] = field(default_factory=dict)
    display_labels: dict[str, dict[str, str]] = field(default_factory=dict)
    row_order: dict[str, dict[str, int]] = field(default_factory=dict)
    skipped_reason: str | None = None


@dataclass
class ConversionResult:
    output_paths: list[Path]
    summaries: dict[str, dict[str, dict[str, float | None]]]
    skipped: list[str]


def convert_pdfs(pdf_paths: Iterable[Path], output_dir: Path, use_ai: bool = False) -> ConversionResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    parsed = [parse_pdf(Path(path), use_ai=use_ai) for path in pdf_paths]

    valid = [item for item in parsed if not item.skipped_reason]
    skipped = [
        f"{item.source_name}: {item.skipped_reason}" for item in parsed if item.skipped_reason
    ]
    grouped: dict[str, list[ParsedPdf]] = {}
    for item in valid:
        grouped.setdefault(item.company, []).append(item)

    output_paths: list[Path] = []
    summaries: dict[str, dict[str, dict[str, float | None]]] = {}

    for company, filings in grouped.items():
        filings.sort(key=lambda item: _period_sort_key(item.periods))
        company_periods = _combined_periods(filings)
        # safe_filename the period tag too — a generic fallback like "Period 1"
        # contains a space, which the download route's secure_filename() would
        # rewrite to "_", breaking the download lookup.
        period_tag = safe_filename(company_periods[-1]) if company_periods else "Output"
        out_name = (
            f"{safe_filename(company)}_Consolidated_FS_MultiYear.xlsx"
            if len(company_periods) > 2 or len(filings) > 1
            else f"{safe_filename(company)}_Consolidated_FS_{period_tag}.xlsx"
        )
        out_path = output_dir / out_name
        build_company_workbook(company, filings, out_path)
        output_paths.append(out_path)
        summaries[company] = build_summary(company, filings)

    if len(grouped) >= 2:  # a master only makes sense across 2+ distinct companies
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        master_path = output_dir / f"MASTER_Consolidated_FS_{timestamp}.xlsx"
        build_master_workbook(grouped, master_path)
        output_paths.append(master_path)

    return ConversionResult(output_paths=output_paths, summaries=summaries, skipped=skipped)


# PDFs with more pages than this use the low-memory two-phase reader, so a big
# annual report (e.g. a 90-page US filing) doesn't run pdfminer over every page
# and get OOM-killed on a small instance. Smaller reports keep the original,
# proven full-pdfplumber path unchanged.
LARGE_PDF_PAGES = 30


def _pdf_page_count(path: Path) -> int:
    import pypdfium2 as pdfium

    pdf = pdfium.PdfDocument(str(path))
    try:
        return len(pdf)
    finally:
        pdf.close()


def _scan_text_pdfium(path: Path) -> list[tuple[int, str]]:
    """Cheap, low-memory full-document text scan via pypdfium2 (C-backed; each
    page is read and released immediately). Used only to locate statement pages
    and the company name in large PDFs — not for precise number parsing."""
    import pypdfium2 as pdfium

    pdf = pdfium.PdfDocument(str(path))
    out: list[tuple[int, str]] = []
    try:
        for i in range(len(pdf)):
            page = pdf.get_page(i)
            textpage = page.get_textpage()
            try:
                text = textpage.get_text_range() or ""
            finally:
                textpage.close()
                page.close()
            out.append((i + 1, text))
    finally:
        pdf.close()
    return out


def _extract_pages_pdfplumber(path: Path, page_numbers: set[int]) -> dict[int, str]:
    """Run pdfplumber's precise extraction on ONLY the given pages, flushing each
    page's cache so peak memory stays bounded regardless of document size."""
    result: dict[int, str] = {}
    if not page_numbers:
        return result
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            page_no = i + 1
            if page_no in page_numbers:
                result[page_no] = _extract_page_columns(page)
            page.flush_cache()
    return result


def _extract_page_columns(page) -> str:
    """Extract a page's text, handling the side-by-side two-statement layout common
    in Indian reports (e.g. Balance Sheet on the left half, Statement of Profit and
    Loss on the right half). A plain top-to-bottom read interleaves the two columns
    ("Right-of-Use Asset" + "II Expenses" -> garbage); when we detect two statement
    titles sitting side by side, we crop the page at the gutter and read the left
    column fully, then the right column fully."""
    base = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
    # Cheap gate: only statement-ish pages are worth the geometry work.
    low = base.lower()
    if not any(s in low for s in ("balance sheet", "profit and loss", "cash flow",
                                  "statement of operations", "statement of income",
                                  "comprehensive income", "financial position")):
        return base
    try:
        words = page.extract_words(x_tolerance=1, y_tolerance=3)
    except Exception:
        return base
    if len(words) < 30:
        return base
    width = float(page.width)
    # Find a clean vertical GUTTER in the central band that no word crosses, with
    # numeric content AND row-start labels on BOTH sides — the signature of two
    # statements printed side by side (Balance Sheet | Profit & Loss).
    has_digit = re.compile(r"\d")
    has_alpha = re.compile(r"[A-Za-z]")
    split_x = None
    lo, hi = width * 0.40, width * 0.60
    steps = 24
    for k in range(steps + 1):
        sx = lo + k * (hi - lo) / steps
        if any(float(w["x0"]) < sx < float(w["x1"]) for w in words):
            continue  # a word crosses here -> not a gutter
        left_nums = sum(1 for w in words if float(w["x1"]) <= sx and has_digit.search(w["text"]))
        right_nums = sum(1 for w in words if float(w["x0"]) >= sx and has_digit.search(w["text"]))
        # the right table must start its own labels just past the gutter
        right_labels = sum(
            1 for w in words
            if sx <= float(w["x0"]) < sx + width * 0.18 and has_alpha.search(w["text"])
        )
        if left_nums >= 6 and right_nums >= 6 and right_labels >= 4:
            split_x = sx
            break
    if split_x is None:
        return base  # single column / stacked — the line-span logic handles it
    left = page.crop((0, 0, split_x, page.height)).extract_text(x_tolerance=1, y_tolerance=3) or ""
    right = page.crop((split_x, 0, width, page.height)).extract_text(x_tolerance=1, y_tolerance=3) or ""
    return left + "\n" + right


def parse_pdf(path: Path, use_ai: bool = False) -> ParsedPdf:
    try:
        page_count = _pdf_page_count(path)
    except Exception:
        page_count = 0  # fall through to the pdfplumber path, which surfaces a real error

    if page_count > LARGE_PDF_PAGES:
        # Low-memory path: cheap full scan to find the statement pages, then
        # deep-parse only those few pages with pdfplumber for accurate numbers.
        try:
            scan_pages = _scan_text_pdfium(path)
            scan_spans = find_statement_pages(scan_pages)
            comp_page = _find_comprehensive_page(scan_pages)
            # Deep-parse the chosen statement pages plus their immediate neighbours
            # (continuation pages, combined-statement neighbours) and the
            # comprehensive-income page, so the final text is precise everywhere
            # the numbers come from.
            needed: set[int] = set()
            for spans in scan_spans.values():
                for (p, _s, _e) in spans:
                    needed.update({p - 1, p, p + 1})
            if comp_page:
                needed.add(comp_page)
            needed = {p for p in needed if p >= 1}
            precise = _extract_pages_pdfplumber(path, needed)
        except Exception as exc:
            return ParsedPdf("Unknown Company", path.name, path, skipped_reason=f"could not read PDF ({exc})")
        # Precise text on statement pages; cheap scan text elsewhere.
        pages = [(page_no, precise.get(page_no, text)) for page_no, text in scan_pages]
        # Keep the scan's (correct) page+statement choices, but recompute the span
        # line-indices on the precise text (cheap scan can paginate differently).
        # NOTE: do NOT re-run the full selector here — it would re-pick the cluster
        # on mixed precise/scan text and can land on the wrong one.
        statement_pages = _spans_on_precise(scan_spans, dict(pages))
    else:
        try:
            with pdfplumber.open(path) as pdf:
                pages = []
                for i, page in enumerate(pdf.pages):
                    text = _extract_page_columns(page)
                    pages.append((i + 1, text))
                    # Release pdfplumber's cached layout objects for this page.
                    page.flush_cache()
        except Exception as exc:
            return ParsedPdf("Unknown Company", path.name, path, skipped_reason=f"could not read PDF ({exc})")
        statement_pages = find_statement_pages(pages)
        comp_page = _find_comprehensive_page(pages)

    page_text = dict(pages)
    # Company detection reads the WHOLE chosen statement pages (their header
    # carries "<Company> and Subsidiaries"), not the sliced spans — the span
    # starts at the statement title, below that header line.
    chosen_pages = sorted({p for spans in statement_pages.values() for (p, _s, _e) in spans})
    statement_preview = "\n".join(page_text.get(p, "") for p in chosen_pages)
    joined_first_pages = "\n".join(text for _, text in pages[:12])
    company = detect_company(statement_preview) or detect_company(joined_first_pages) or path.stem

    if not any(statement_pages.values()):
        return ParsedPdf(company, path.name, path, skipped_reason="no detectable consolidated financial statements")

    parsed = ParsedPdf(company=company, source_name=path.name, source_path=path)
    parsed.statements = {}

    all_periods: list[str] = []
    for statement_name, spans in statement_pages.items():
        statement_text = _span_text(page_text, spans)
        periods = detect_periods(statement_text, statement_name)
        if not periods:
            periods = detect_periods(joined_first_pages + "\n" + statement_text, statement_name)
        if periods:
            all_periods.extend(periods)
        rows, labels, order = parse_statement_rows(statement_text, periods, statement_name)
        if rows:
            parsed.statements[statement_name] = rows
            parsed.display_labels[statement_name] = labels
            parsed.row_order[statement_name] = order

    # Fold the separate Comprehensive Income statement's two summary lines into
    # the Income Statement's Comprehensive Income section.
    if comp_page is not None and "Income Statement" in parsed.statements:
        comp_text = next((t for pn, t in pages if pn == comp_page), "")
        if comp_text:
            cperiods = detect_periods(comp_text, "Income Statement") or _dedupe(all_periods)
            crows, clabels, corder = parse_statement_rows(comp_text, cperiods, "Income Statement")
            is_rows = parsed.statements["Income Statement"]
            is_labels = parsed.display_labels["Income Statement"]
            is_order = parsed.row_order["Income Statement"]
            base = (max(is_order.values()) if is_order else 0) + 1000
            for key in crows:
                if "comprehensive" in clabels.get(key, "").lower():
                    is_rows[key] = crows[key]
                    is_labels[key] = clabels[key]
                    is_order[key] = base + corder.get(key, 0)

    parsed.periods = _dedupe_periods(all_periods)
    if not parsed.periods:
        parsed.periods = _infer_periods_from_rows(parsed.statements)

    # AI-assisted extraction: when the user enables it, the LLM does the FULL
    # structuring (the way the manual "A2E" process does), replacing the
    # deterministic rows with its clean line items. This is the path to top
    # quality on tricky / dense reports.
    if use_ai:
        # Hand the LLM the WHOLE statement-region pages (not the narrow detected
        # spans) so it can find statements the detector mis-assigned and read
        # messy / multi-column / digit-split tables itself — the A2E approach.
        # Cap the page span so a stray detection can't balloon the request.
        span_pages = sorted({p for spans in statement_pages.values() for (p, _s, _e) in spans})
        if comp_page:
            span_pages = sorted(set(span_pages + [comp_page]))
        if span_pages and (span_pages[-1] - span_pages[0]) <= 12:
            lo, hi = span_pages[0] - 1, span_pages[-1] + 1
            region = [page_text[p] for p in range(lo, hi + 1) if p in page_text]
        else:  # detections too far apart — fall back to the per-statement spans
            region = [_span_text(page_text, s) for s in statement_pages.values() if s]
        blob = "\n\n".join(t for t in region if t)
        if blob.strip():
            _apply_ai_full(parsed, {"Statements": blob})

    if not parsed.statements:
        parsed.skipped_reason = "consolidated statement pages found but no rows could be parsed"

    return parsed


def _parse_confidence(parsed: ParsedPdf) -> int:
    """How many of the 3 statements parsed with a usable number of rows.
    3 = all good; lower means the deterministic pass likely missed something."""
    good = 0
    for statement in STATEMENTS:
        rows = parsed.statements.get(statement, {})
        if len(rows) >= 4:
            good += 1
    return good


def _apply_ai_fallback(parsed: ParsedPdf, statement_texts: dict[str, str]) -> None:
    """Merge Gemini-extracted rows into `parsed`, filling statements/lines the
    deterministic parser missed. Never overwrites already-parsed rows. No-op if
    AI isn't configured or returns nothing."""
    try:
        import gemini_fallback
    except Exception:
        return
    ai = gemini_fallback.extract(statement_texts, parsed.periods)
    if not ai:
        return
    ai_periods = ai.get("periods") or []
    for period in ai_periods:
        if period not in parsed.periods:
            parsed.periods.append(period)
    for statement, payload in ai.get("statements", {}).items():
        rows = parsed.statements.setdefault(statement, {})
        labels = parsed.display_labels.setdefault(statement, {})
        order = parsed.row_order.setdefault(statement, {})
        base = (max(order.values()) if order else 0) + 1
        for key, values in payload.get("rows", {}).items():
            if key in rows:  # trust the on-server deterministic value first
                continue
            rows[key] = values
            labels[key] = payload.get("labels", {}).get(key, key.title())
            order[key] = base + payload.get("order", {}).get(key, 0)
    parsed.periods = _dedupe_periods(parsed.periods)


def _apply_ai_full(parsed: ParsedPdf, statement_texts: dict[str, str]) -> None:
    """REPLACE each statement with the LLM's full clean extraction (the A2E-style
    path). Falls back to the deterministic rows for any statement the AI returns
    nothing for. No-op if AI isn't configured or returns nothing at all."""
    try:
        import gemini_fallback
    except Exception:
        return
    ai = gemini_fallback.extract(statement_texts, parsed.periods)
    if not ai:
        return  # keep the deterministic result if AI unavailable/failed
    ai_periods = [p for p in (ai.get("periods") or []) if p]
    if ai_periods:
        parsed.periods = _dedupe_periods(ai_periods + parsed.periods)
    for statement, payload in ai.get("statements", {}).items():
        rows = payload.get("rows", {})
        if not rows:  # AI gave nothing for this statement -> keep deterministic
            continue
        parsed.statements[statement] = rows
        parsed.display_labels[statement] = payload.get("labels", {})
        parsed.row_order[statement] = payload.get("order", {})


def detect_company(text: str) -> str | None:
    normalized = normalize_text(text)
    annual_report_match = re.search(
        r"([A-Z][A-Za-z0-9&.,'’() -]{3,}?)\s+Annual Report\s+20\d{2}",
        normalized,
        flags=re.I,
    )
    if annual_report_match:
        return title_company(annual_report_match.group(1).strip(" -.,"))

    members_match = re.search(
        r"To the Members of\s+([A-Z][A-Za-z0-9&.,'’() -]{3,}?)(?:\s+Report on|\s+Basis for|\n)",
        normalized,
        flags=re.I,
    )
    if members_match:
        return title_company(members_match.group(1).strip(" -.,"))

    result_match = re.search(
        r"^([A-Z][A-Z0-9&.,'’() -]{3,}?(?:LIMITED|LTD\.?|INC\.?|CORPORATION|COMPANY|PLC))\s*$",
        normalized,
        flags=re.I | re.M,
    )
    if result_match:
        return title_company(result_match.group(1).strip(" -.,"))

    patterns = [
        r"([A-Z][A-Z0-9&.,'’() -]{3,}?(?:LIMITED|LTD\.?|INC\.?|CORPORATION|COMPANY|PLC))",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized, flags=re.I)
        if match:
            candidate = re.sub(r"\s+", " ", match.group(1)).strip(" -.,")
            candidate = re.sub(r"\bAnnual Report\b.*", "", candidate, flags=re.I).strip()
            if len(candidate) > 3:
                return title_company(candidate)
    return None


def detect_company(text: str) -> str | None:
    normalized = normalize_text(text)
    for line in normalized.splitlines():
        annual_report_match = re.search(
            r"^([A-Z][A-Za-z0-9&.,'() -]{3,}?)\s+Annual Report\s+20\d{2}",
            line.strip(),
            flags=re.I,
        )
        if annual_report_match:
            return title_company(annual_report_match.group(1).strip(" -.,"))

        # A standalone company line, e.g. "BRITANNIA INDUSTRIES LIMITED" at the
        # top of a SEBI quarterly results page. \b after the suffix stops "income"
        # from matching "Inc".
        exact_company = re.search(
            r"^([A-Z][A-Z0-9&.,'() -]{3,}?(?:LIMITED|LTD\.?|INC\.?|CORPORATION|COMPANY|PLC))\b\s*$",
            line.strip(),
            flags=re.I,
        )
        if exact_company:
            return title_company(exact_company.group(1).strip(" -.,"))

    members_match = re.search(
        r"To the Members of\s+([A-Z][A-Za-z0-9&.,'() -]{3,}?)(?:\s+Report on|\s+Basis for|\n)",
        normalized,
        flags=re.I,
    )
    if members_match:
        return title_company(members_match.group(1).strip(" -.,"))

    # \b after the suffix prevents matching "Interest inc-ome", "...company-wide",
    # etc. — the suffix must be a whole word.
    generic_match = re.search(
        r"([A-Z][A-Za-z0-9&.,'() -]{3,}?(?:LIMITED|LTD\.?|INC\.?|CORPORATION|COMPANY|PLC))\b",
        normalized,
        flags=re.I,
    )
    if generic_match:
        candidate = generic_match.group(1).strip(" -.,")
        # Guard against a lowercase-prose false positive ("... the company")
        # by requiring the matched span to contain an uppercase anchor.
        if re.search(r"[A-Z]{2,}", candidate) or candidate.istitle():
            return title_company(candidate)
    return None


# Title phrases that, when they START a heading line, identify a primary
# consolidated statement page (covers both US 10-K and Indian annual-report
# wording).
_TITLE_PATTERNS = {
    "Income Statement": (
        r"consolidated statements? of operations",
        r"consolidated statements? of income",
        r"consolidated statement of profit and loss",
        r"consolidated statement of profit",
        r"statement of audited consolidated financial results",
    ),
    "Balance Sheet": (
        r"consolidated balance sheets?",
    ),
    "Cash Flow Statement": (
        r"consolidated statements? of cash flows?",
        r"consolidated cash flows? statement",
        r"consolidated cash flow statements?",
    ),
}

# Pages that mention the statement titles only in prose / listings — never the
# primary statement page itself.
_PAGE_EXCLUDE_MARKERS = (
    "index to consolidated financial statements",
    "report of independent registered public accounting firm",
    "we have audited",
    "item 8. financial statements",
    "item 15",
    "part iv",
)

_PERIOD_HEADER_RE = re.compile(
    r"(year ended|quarter ended|months ended|period ended|as at|as of|"
    r"(?:january|february|march|april|may|june|july|august|september|october|november|december)"
    r"\s+\d{1,2}|\b20\d{2}\b)",
    re.I,
)


def _title_heading_score(line: str, patterns: tuple[str, ...]) -> int:
    """Score a line as a statement-title heading. A title that starts the line
    and is followed by nothing / a date / "(in millions)" scores high; an index
    entry ("... 38") or a mid-sentence mention scores 0."""
    stripped = re.sub(r"\s+", " ", line).strip().lower().rstrip(".")
    if not stripped or len(stripped) > 90:
        return 0
    # 10-Q / interim statements are titled "Condensed Consolidated Statements of
    # Income" etc. Drop a leading qualifier so the title still matches.
    stripped = re.sub(r"^(?:unaudited\s+|condensed\s+|interim\s+)+", "", stripped)
    for pat in patterns:
        match = re.match(rf"{pat}\b", stripped)
        if not match:
            continue
        rest = stripped[match.end():].strip(" .,:-")
        if re.fullmatch(r"\d{1,4}", rest):  # index entry: title + page number
            return 0
        if rest == "" or rest.startswith("("):
            return 3
        if len(rest) <= 40 and _PERIOD_HEADER_RE.search(rest):  # "... as at March 31, 2023"
            return 3
        return 0  # title trails into a sentence -> notes / prose
    return 0


def _numeric_line_count(text: str) -> int:
    count = 0
    for line in text.splitlines():
        if len(re.findall(r"\d", line)) >= 3 and re.search(r"\d[\d,]*", line):
            count += 1
    return count


_COMPREHENSIVE_PATTERNS = (r"consolidated statements? of comprehensive income",)


def _find_comprehensive_page(pages: list[tuple[int, str]]) -> int | None:
    """The standalone 'Consolidated Statements of Comprehensive Income' page
    (a separate statement in US 10-Ks), whose two summary lines we fold into the
    Income Statement's Comprehensive Income section."""
    best_score, best_page = 0, None
    for page_no, text in pages:
        if any(marker in text.lower() for marker in _PAGE_EXCLUDE_MARKERS):
            continue
        nlines = _numeric_line_count(text)
        if nlines < 3:
            continue
        head_lines = [ln for ln in text.splitlines() if ln.strip()][:12]
        heading = max(
            (_title_heading_score(line, _COMPREHENSIVE_PATTERNS) for line in head_lines),
            default=0,
        )
        if heading > 0 and heading * 100 + nlines > best_score:
            best_score, best_page = heading * 100 + nlines, page_no
    return best_page


# A span is (page_no, start_line, end_line) into that page's text — so a page that
# packs two statements (e.g. an Indian "Balance Sheet" + "Statement of Profit and
# Loss" on one sheet) can be split between them.
Span = tuple[int, int, int]


def find_statement_pages(pages: list[tuple[int, str]]) -> dict[str, list[Span]]:
    page_text = {pno: txt for pno, txt in pages}

    # 1) Collect every clean statement-title heading ANYWHERE on each dense page
    #    (not just the top), with its line index — this catches combined pages.
    candidates: dict[str, list[tuple]] = {name: [] for name in STATEMENTS}
    titles_on_page: dict[int, list[tuple[int, str]]] = {}
    for page_no, text in pages:
        lower = text.lower()
        if any(marker in lower for marker in _PAGE_EXCLUDE_MARKERS):
            continue
        numeric_lines = _numeric_line_count(text)
        if numeric_lines < 5:  # a real statement page is dense with numbers
            continue
        lines = text.splitlines()
        page_titles: list[tuple[int, str]] = []
        for i, line in enumerate(lines):
            if not line.strip():
                continue
            for name, patterns in _TITLE_PATTERNS.items():
                heading = _title_heading_score(line, patterns)
                if heading <= 0:
                    continue
                is_cons = "consolidated" in line.lower()
                score = heading * 100 + min(numeric_lines, 60) + (60 if is_cons else 0)
                candidates[name].append((page_no, i, score, is_cons))
                page_titles.append((i, name))
                break
        if page_titles:
            titles_on_page[page_no] = sorted(page_titles)

    if not any(candidates.values()):
        return {name: [] for name in STATEMENTS}

    # 2) Indian reports carry BOTH standalone and consolidated statements; pick the
    #    CONSOLIDATED set. More generally, find the densest cluster — the anchor
    #    page near which the most distinct statements have a candidate (US 10-K
    #    statements span a few pages; Indian ones sit together / on one page).
    doc_has_consolidated = any(c[3] for cl in candidates.values() for c in cl)
    all_pages = sorted({c[0] for cl in candidates.values() for c in cl})

    def cluster_quality(anchor: int) -> tuple:
        lo, hi = anchor - 3, anchor + 6
        names, total, cons = set(), 0, 0
        for name, cl in candidates.items():
            near = [c for c in cl if lo <= c[0] <= hi and (c[3] or not doc_has_consolidated)]
            if near:
                names.add(name)
                best = max(near, key=lambda c: c[2])
                total += best[2]
                cons += 1 if best[3] else 0
        return (len(names), cons, total)

    anchor = max(all_pages, key=cluster_quality)
    lo, hi = anchor - 3, anchor + 6

    # 3) Assign each statement its best candidate inside the cluster window,
    #    preferring consolidated; fall back to global best if none nearby.
    chosen: dict[str, tuple[int, int]] = {}
    for name, cl in candidates.items():
        near = [c for c in cl if lo <= c[0] <= hi]
        pool = near or cl
        if doc_has_consolidated and any(c[3] for c in pool):
            pool = [c for c in pool if c[3]]
        if pool:
            best = max(pool, key=lambda c: c[2])
            chosen[name] = (best[0], best[1])

    # 4) Build spans, ending each statement at the NEXT title on its page (so a
    #    combined Balance-Sheet + P&L page is split cleanly between them).
    spans: dict[str, list[Span]] = {name: [] for name in STATEMENTS}
    for name, (page_no, line_idx) in chosen.items():
        n_lines = len(page_text[page_no].splitlines())
        # End at the next DIFFERENT statement's title (a repeated same-statement
        # title is just a header reprint, e.g. Paytm prints the P&L title twice).
        following = [li for li, nm in titles_on_page.get(page_no, []) if li > line_idx and nm != name]
        end = min(following) if following else n_lines
        spans[name].append((page_no, line_idx, end))

    # 5) Continuation: a statement that runs to the bottom of its page can spill
    #    onto the next (cash-flow financing tail, balance-sheet equity half).
    for name in ("Cash Flow Statement", "Balance Sheet"):
        for (page_no, _start, end) in list(spans[name]):
            if end < len(page_text[page_no].splitlines()):
                continue  # ended mid-page (a combined page) -> no spill
            nxt = page_text.get(page_no + 1, "")
            if not nxt or _numeric_line_count(nxt) < 5:
                continue
            if page_no + 1 in titles_on_page:  # next page begins a new statement
                continue
            low = nxt.lower()
            if name == "Cash Flow Statement":
                keys = ("financing activities", "net cash", "cash and cash equivalents", "continued")
                if any(k in low for k in keys):
                    spans[name].append((page_no + 1, 0, len(nxt.splitlines())))
            else:
                if "stockholders" in low and "equity" in low and "statement" in low:
                    continue
                keys = ("total equity and liabilities", "total liabilities and equity", "continued")
                if any(k in low for k in keys):
                    spans[name].append((page_no + 1, 0, len(nxt.splitlines())))
    return spans


def _span_text(page_text: dict[int, str], spans: list[Span]) -> str:
    """Concatenate the line-ranges for a statement's spans into one text block."""
    parts = []
    for (page_no, start, end) in spans:
        lines = page_text.get(page_no, "").splitlines()
        parts.append("\n".join(lines[start:end]))
    return "\n".join(parts)


def _titles_on(text: str) -> list[tuple[int, str]]:
    """Line index + statement name for every clean statement-title heading in a
    page's text (used to recompute spans on precise text)."""
    out: list[tuple[int, str]] = []
    for i, line in enumerate(text.splitlines()):
        if not line.strip():
            continue
        for name, patterns in _TITLE_PATTERNS.items():
            if _title_heading_score(line, patterns) > 0:
                out.append((i, name))
                break
    return sorted(out)


def _spans_on_precise(scan_spans: dict[str, list[Span]], page_text: dict[int, str]) -> dict[str, list[Span]]:
    """Re-locate each scan-chosen span on the precise (pdfplumber) text, keeping
    the scan's page+statement decision but fixing line indices to the precise
    pagination. Splits a combined page at the next different-statement title."""
    titles = {p: _titles_on(page_text.get(p, "")) for sp in scan_spans.values() for (p, _s, _e) in sp}
    out: dict[str, list[Span]] = {name: [] for name in STATEMENTS}
    for name, spans in scan_spans.items():
        for (page_no, s_scan, _e_scan) in spans:
            n_lines = len(page_text.get(page_no, "").splitlines())
            page_titles = titles.get(page_no, [])
            same = [li for li, nm in page_titles if nm == name]
            if same:  # title page: start at this statement's title (nearest scan pos)
                start = min(same, key=lambda li: abs(li - s_scan))
            else:  # continuation page (scan start was 0) or title not re-found
                start = 0
            following = [li for li, nm in page_titles if li > start and nm != name]
            end = min(following) if following else n_lines
            out[name].append((page_no, start, end))
    return out


def detect_periods(text: str, statement_name: str) -> list[str]:
    normalized = normalize_text(text)
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]

    if "quarter ended" in normalized.lower() and "year ended" in normalized.lower():
        dated_years = re.findall(r"\b\d{1,2}[./-]\d{1,2}[./-](20\d{2})\b", normalized)
        if len(dated_years) >= 2:
            return [f"FY{year}" for year in _dedupe(dated_years[-2:])]

    # Month names AND abbreviations (NVIDIA's FY ends "Jan 25, 2026"; Indian
    # reports use "March 31, 2025"). "sept" tolerated.
    months = (r"(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*")
    header = lines[:16]  # the statement title + column-header block

    # 1) Years embedded directly in a dated column header — "<Month> <day>, <year>"
    #    or "<day> <Month> <year>" or "31/03/2025". Collect across the header block
    #    in reading order (left->right / top->bottom matches the value columns).
    years: list[str] = []
    for line in header:
        low = line.lower()
        for m in re.finditer(rf"{months}\.?\s+\d{{1,2}},?\s*(20\d{{2}})", low):
            years.append(m.group(1))
        for m in re.finditer(rf"\d{{1,2}}\s+{months}\.?\s*,?\s*(20\d{{2}})", low):
            years.append(m.group(1))
        for m in re.finditer(r"\b\d{1,2}[./-]\d{1,2}[./-](20\d{2})\b", low):
            years.append(m.group(1))
    years = _dedupe(years)

    # 2) US style where the date label and the year row are SEPARATE lines
    #    ("Year Ended December 31," then "2023 2024 2025"). Grab years from a small
    #    window anchored on the date/period phrase.
    if len(years) < 2:
        for idx, line in enumerate(header):
            low = line.lower()
            if ("year ended" in low or "as at" in low or "as of" in low
                    or "december 31" in low or re.search(rf"{months}\.?\s+\d{{1,2}}", low)):
                window = " ".join(header[idx : idx + 3])
                w = _dedupe(re.findall(r"\b(20\d{2})\b", window))
                if len(w) > len(years):
                    years = w
                if len(years) >= 2:
                    break

    # 3) A standalone row of bare years ("2023 2024 2025").
    if len(years) < 2:
        for line in header:
            if re.fullmatch(r"(?:20\d{2}\s*){2,6}", line.strip()):
                years = re.findall(r"20\d{2}", line)
                break

    # 4) Last resort: any single year in the header block.
    if not years:
        for line in header:
            found = re.findall(r"\b20\d{2}\b", line)
            if found:
                years = _dedupe(found)
                break

    if statement_name == "Balance Sheet" and len(years) > 2:
        years = years[:2]  # US/Indian balance sheets show 2 columns (current, prior)
    if len(years) > 6:
        years = years[:6]

    return [f"FY{year}" for year in years]


def parse_statement_rows(
    text: str, periods: list[str], statement_name: str
) -> tuple[dict[str, dict[str, float]], dict[str, str], dict[str, int]]:
    periods = periods or []
    period_count = max(1, len(periods))
    rows: dict[str, dict[str, float]] = {}
    labels: dict[str, str] = {}  # canonical key -> clean display label
    order: dict[str, int] = {}   # canonical key -> first line position (PDF row order)
    current_header = ""          # sub-section header ("Net sales:", "Cost of sales:")

    for idx, raw_line in enumerate(normalize_text(text).splitlines()):
        line = raw_line.strip()
        if not line or len(line) < 4:
            continue
        lower = line.lower()
        if any(skip in lower for skip in ("see accompanying notes", "corporate overview",
                                          "statutory reports", "form 10-k", "form 10k", "table of contents",
                                          "par value", "shares authorized", "shares issued",
                                          "issued and outstanding", "shares outstanding", "respectively")):
            continue  # share-count descriptions carry bogus numbers (authorized shares, dates)

        numbers = extract_numbers(line)
        # A date-column header whose day+year fused into a 6-digit "DDYYYY" token
        # ("March 31, 2025" -> 312025; NVIDIA "Jan 25, 2026" -> 252026). If every
        # number on the line decodes to a valid date, it's a header, not a row.
        if numbers and all(_is_date_encoded(n) for n in numbers):
            continue
        # A sub-section header ("Net sales:", "Cost of sales:", "Operating
        # expenses:") — track it so bare lines under it (Apple "Products" /
        # "Services") get the right context and don't collide.
        if line.endswith(":") and not numbers:
            current_header = re.sub(r"\s+", " ", line.rstrip(":").strip())
            continue
        # A real statement row has exactly the period columns (allow one stray,
        # e.g. a note reference). The statement page is detected precisely, so
        # there are no footnote tables here to capture extra numbers from.
        if not (period_count <= len(numbers) <= period_count + 1):
            continue
        text_part = strip_numbers_from_label(line)
        if not re.search(r"[A-Za-z]", text_part) or len(text_part.strip()) < 3:
            continue
        # Skip the period-header line itself ("Year Ended December 31, 2022 2023"),
        # whose "numbers" are just years/dates, not financial values.
        if re.search(_MONTH_RE, lower) and all(1900 <= abs(n) <= 2100 for n in numbers):
            continue
        # A date-header line whose label is ONLY month name(s) — e.g. Apple's
        # "September 27, 2025  September 28, 2024" (the day+year merge into bogus
        # numbers, dodging the year-range check above).
        if re.search(_MONTH_NAMES, lower) and not re.sub(
            _MONTH_NAMES, "", text_part, flags=re.I
        ).strip(" ,.;:-0123456789$()"):
            continue
        if lower.startswith("year ended") or "months ended" in lower or lower.strip() in ("period", "particulars"):
            continue
        # Indian date-column headers ("For the year ended March 31, 2025",
        # "As at March 31, 2025", "As of December 31,") whose day+year fuse into a
        # bogus number like 312025 — always a header, never a line item.
        if (("for the year ended" in lower or "for the period ended" in lower
                or "for the quarter ended" in lower or lower.startswith("as at")
                or lower.startswith("as of") or lower.startswith("particulars"))
                and re.search(_MONTH_NAMES, lower)):
            continue

        # Drop a single leading note-reference (small integer) when it shows up
        # as one extra number, e.g. Indian "3  1,13,114  1,60,281".
        if len(numbers) == period_count + 1 and abs(numbers[0]) < 100 and numbers[0] == int(numbers[0]):
            numbers = numbers[1:]
        values = numbers[:period_count] if statement_name == "Balance Sheet" else numbers[-period_count:]
        raw_label = re.sub(r"\s+", " ", text_part).strip(" -:;*")
        low_lbl = raw_label.lower()
        # In the income statement a lone "Basic"/"Diluted" line is the
        # weighted-average share count (its parent header wraps to its own line).
        if statement_name == "Income Statement" and low_lbl in ("basic", "diluted"):
            which = "Basic" if low_lbl == "basic" else "Diluted"
            # A bare "Basic"/"Diluted" line is either EPS (small, often decimal)
            # or the weighted-average share count (large whole numbers). Both can
            # appear (e.g. Meta), so tell them apart by the values.
            nums = [v for v in values if v is not None]
            has_decimal = any(abs(v - round(v)) > 1e-9 for v in nums)
            per_share = has_decimal or (nums and max(abs(v) for v in nums) < 100)
            if per_share:
                key = f"{which} EPS"
                display = f"{which} Earnings per Share"
            else:
                key = f"Weighted-Avg Shares {which}"
                display = f"Weighted-Avg Shares ({which})"
        elif statement_name == "Cash Flow Statement" and low_lbl in ("period", "of period"):
            # The "...beginning of period" cash line wraps so only "Period" survives.
            key = "Cash Beginning of Period"
            display = "Cash, Cash Equivalents & Restricted Cash, Beginning of Period"
        elif (low_lbl in ("products", "product", "services", "service", "goods", "equipment", "software")
              and "cost" in current_header.lower()):
            # Apple lists "Products"/"Services" under BOTH "Net sales:" and
            # "Cost of sales:" — keep the cost ones distinct from the revenue ones.
            key = standardize_label(f"Cost of Sales {raw_label}", statement_name)
            display = f"{clean_label(raw_label)} (Cost of Sales)"
        else:
            key = standardize_label(raw_label, statement_name)  # merge key only
            if not key or len(key) < 3:
                continue
            if re.fullmatch(r"[ivxlcdm.\s()-]+", key.lower()):
                continue
            display = clean_label(raw_label)  # polished Title Case for display

        if not periods:
            periods = [f"Period {i + 1}" for i in range(period_count)]
        rows[key] = {period: value for period, value in zip(periods[-len(values) :], values)}
        labels[key] = display
        order.setdefault(key, idx)

    return rows, labels, order


def normalize_text(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = text.replace("`", "₹")
    text = text.replace("’", "'")
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"(?<=\d)\s*,\s*(?=\d)", ",", text)
    text = re.sub(r"(?<=\d)\s*\.\s*(?=\d)", ".", text)
    # NOTE: do NOT collapse "989 2,949" -> "9892,949"; a space before a 4-digit
    # "X,XXX" thousands value is a column break, not a split single number.
    text = re.sub(r"(\.\d{2})(?=\d{1,3},\d{3}\.\d{2})", r"\1 ", text)
    return text


def _is_date_encoded(n: float) -> bool:
    """True if n is a 6-digit 'DDYYYY' date token produced when a column header's
    day and year fuse — e.g. 'March 31, 2025' -> 312025, 'Jan 25, 2026' -> 252026.
    Used to drop date-header lines that masquerade as data rows."""
    if n != int(n):
        return False
    n = abs(int(n))
    if 10100 <= n <= 312099:
        day, year = divmod(n, 10000)
        return 1 <= day <= 31 and 2000 <= year <= 2099
    return False


def extract_numbers(line: str) -> list[float]:
    # A standalone dash is a zero placeholder ("repurchased  —  (6,000)  —").
    line = re.sub(r"(?<=\s)[-–—](?=\s|$)", "0", line)
    matches = re.findall(r"\(?-?\d[\d,]*(?:\.\d+)?\)?", line)
    values: list[float] = []
    for token in matches:
        negative = token.startswith("(") and token.endswith(")")
        clean = token.strip("()").replace(",", "")
        try:
            value = float(clean)
        except ValueError:
            continue
        values.append(-value if negative else value)
    return values


def strip_numbers_from_label(line: str) -> str:
    label = re.sub(r"\(?-?\d[\d,]*(?:\.\d+)?\)?", " ", line)
    label = re.sub(r"^\s*(?:[IVXLCDM]+|[A-Z]|\(?[a-z]\)?|S\.?No\.?)\s+", "", label, flags=re.I)
    label = re.sub(r"\s+", " ", label)
    label = label.replace("$", "").replace("₹", "").strip(" :-")
    return label


def standardize_label(label: str, statement_name: str) -> str:
    label = re.sub(r"\bNote\b", "", label, flags=re.I)
    label = re.sub(r"\bRefer note\b", "", label, flags=re.I)
    label = re.sub(r"\s+", " ", label).strip(" -:;")
    lower = label.lower()

    if statement_name == "Balance Sheet":
        # Grand total ("Total liabilities and stockholders' equity") must be
        # caught before the bare "total liabilities" check, or it gets mislabeled
        # as Total Liabilities with the wrong (= total assets) value.
        if "total liabilities and" in lower or "total equity and liabilities" in lower:
            return "Total Equity and Liabilities"
        if "total current assets" in lower:
            return "Total Current Assets"
        if "total current liabilities" in lower:
            return "Total Current Liabilities"
        if "total assets" in lower:
            return "Total Assets"
        if "total stockholders" in lower or "total shareholders" in lower or "total equity" in lower:
            return "Total Equity"
        if "total liabilities" in lower:
            return "Total Liabilities"
    if statement_name == "Income Statement":
        if "sale of goods" in lower:
            return "Sale of Goods"
        if "other operating revenues" in lower:
            return "Other Operating Revenues"
        # Canonicalize across year-to-year wording changes so the same concept
        # is ONE row (e.g. "Income (loss) before income taxes" == "Income before
        # income taxes"; "Net income (loss)" == "Net income").
        # Pre-tax income — "before" + "income tax" covers "Income (loss) before
        # income taxes" AND "Income before provision for income taxes" (Meta).
        if ("income tax" in lower and "before" in lower) or "profit before tax" in lower:
            return "Income Before Income Taxes"
        # The tax charge itself — must NOT contain "before" (else it would
        # swallow the pre-tax line above).
        if ("income tax" in lower and "before" not in lower
                and ("provision" in lower or "benefit" in lower or "tax expense" in lower)):
            return "Provision for Income Taxes"
        if "technology and" in lower and ("content" in lower or "infrastructure" in lower):
            return "Technology and Infrastructure"
        if "total other comprehensive" in lower or ("other comprehensive" in lower and "total" in lower):
            return "Total Other Comprehensive Income"
        if "comprehensive income" in lower:
            return "Comprehensive Income"
        if (
            "profit for the year" in lower
            or "net income" in lower
            or "net profit" in lower
            or "profit/(loss) for the" in lower
        ):
            return "Profit for the Year"

    if statement_name == "Cash Flow Statement":
        # Wording of this line drifts year to year ("... and other" vs
        # "... non-marketable investments, and other"); collapse to one row.
        if "acquisitions" in lower and "net of cash acquired" in lower:
            return "Acquisitions, Net of Cash Acquired, and Other"
        # The reconciling adjustment for "Other income (expense), net" is reworded
        # across years ("Other operating expense (income), net", "Non-operating
        # expense (income), net" …) — collapse to one operating-section row.
        if "expense (income), net" in lower:
            return "Other Expense (Income), Net (Adj.)"

    replacements = {
        "total revenue from operations": "Revenue from Operations",
        "total net sales": "Total Net Sales",
        "net income": "Net Income",
        "profit before tax": "Profit Before Tax",
        "income before income taxes": "Income Before Income Taxes",
        "profit for the year": "Profit for the Year",
        "net profit for the period / year": "Profit for the Year",
        "total assets": "Total Assets",
        "total equity": "Total Equity",
        "total liabilities": "Total Liabilities",
        "net cash provided by operating activities": "Net Cash Provided by Operating Activities",
        "net cash generated from operating activities": "Net Cash Generated from Operating Activities",
        "net cash provided by (used in) operating activities": "Net Cash Provided by Operating Activities",
        "net cash used in investing activities": "Net Cash Used in Investing Activities",
        "net cash used in financing activities": "Net Cash Used in Financing Activities",
    }
    for needle, replacement in replacements.items():
        if needle in lower:
            return replacement
    return title_label(label)


def title_company(value: str) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    # Indian reports place the auditor firm + "Chartered Accountants" immediately
    # before the company name in the signature/audit block — keep only the part
    # after the LAST "Chartered Accountants".
    matches = list(re.finditer(r"chartered accountants", value, flags=re.I))
    if matches:
        value = value[matches[-1].end():].strip(" -.,")
    # Drop a leading "for and on behalf of the Board of Directors of" connective.
    value = re.sub(
        r"^for\s+and\s+on\s+behalf\s+of\s+the\s+board\s+of\s+directors\s+of\s+",
        "", value, flags=re.I,
    ).strip()
    # Trim any leading non-letter noise (stray numbers/parens that the greedy
    # company-suffix regex swept in, e.g. Techno's "(1,64,049.12) 20,454.15 ...").
    value = re.sub(r"^[^A-Za-z]+", "", value).strip()
    if value.isupper():
        value = value.title()
    value = value.replace("Amazon.Com", "Amazon.com")
    return value


_SMALL_WORDS = {"of", "and", "the", "for", "to", "in", "on", "a", "an", "or",
                "per", "by", "with", "from", "as", "at"}
_MONTH_NAMES = (
    r"(?:january|february|march|april|may|june|july|august|september|october|"
    r"november|december)"
)
_MONTH_RE = _MONTH_NAMES + r"\s+\d{1,2}"


def clean_label(label: str) -> str:
    """Polished finance Title Case for display, e.g. 'net product sales' ->
    'Net Product Sales', 'cost of sales' -> 'Cost of Sales'. Keeps small words
    lowercase (except first), preserves hyphen casing and parenthetical words."""
    label = re.sub(r"\s+", " ", label).strip(" -:;*,.")
    if not label:
        return label

    def fix(token: str, first: bool) -> str:
        out = []
        for piece in re.split(r"(-)", token):
            if piece == "-":
                out.append(piece)
                continue
            match = re.match(r"^([^A-Za-z]*)(.*)$", piece)
            lead, rest = match.group(1), match.group(2)
            if rest:
                low = rest.lower()
                if not first and low in _SMALL_WORDS:
                    rest = low
                else:
                    rest = rest[0].upper() + rest[1:].lower()
            out.append(lead + rest)
            first = False
        return "".join(out)

    words = label.split(" ")
    return " ".join(fix(w, i == 0) for i, w in enumerate(words))


def title_label(value: str) -> str:
    keep_upper = {"EPS", "PBT", "PAT", "OCI", "AWS"}
    words = []
    for word in value.split():
        clean = word.strip()
        if clean.upper() in keep_upper:
            words.append(clean.upper())
        elif clean.isupper() and len(clean) <= 4:
            words.append(clean)
        else:
            words.append(clean[:1].upper() + clean[1:])
    return " ".join(words)


def build_company_workbook(company: str, filings: list[ParsedPdf], out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    periods = _combined_periods(filings)
    for statement in STATEMENTS:
        ws = wb.create_sheet(statement)
        write_statement_sheet(ws, company, filings, statement, periods)
    wb.save(out_path)


def build_master_workbook(grouped: dict[str, list[ParsedPdf]], out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for statement in STATEMENTS:
        ws = wb.create_sheet(statement)
        write_master_sheet(ws, grouped, statement)
    wb.save(out_path)


# ----- analyst-grade statement styling (ported from the A2E layout) -----
_AS_HEAD = PatternFill("solid", fgColor="1B3A5C")
_AS_SUBHEAD = PatternFill("solid", fgColor="FF9900")
_AS_SECTION = PatternFill("solid", fgColor="FFF3E0")
_AS_TOTAL = PatternFill("solid", fgColor="FFE0B2")
_AS_ALT = PatternFill("solid", fgColor="FAFAFA")
_AS_WHITE = PatternFill("solid", fgColor="FFFFFF")
_AS_NOTE = PatternFill("solid", fgColor="FFF9F0")
_AS_YEAR = PatternFill("solid", fgColor="B35900")

_AS_TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
_AS_SUB_FONT = Font(name="Calibri", bold=True, color="1B3A5C", size=10)
_AS_YEAR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_AS_SEC_FONT = Font(name="Calibri", bold=True, color="1B3A5C", size=10)
_AS_NORM_FONT = Font(name="Calibri", color="000000", size=10)
_AS_NOTE_FONT = Font(name="Calibri", italic=True, color="595959", size=9)
_AS_THIN = Side(style="thin", color="FFB74D")
_AS_MED = Side(style="medium", color="FF9900")

# Ordered sections per statement and the label keywords that fall under each.
_SECTIONS: dict[str, list[tuple[str, tuple[str, ...]]]] = {
    "Income Statement": [
        ("I.  REVENUE", ("revenue from operations", "total revenue", "net revenue", "revenue", "turnover",
                         "product sales", "service sales", "sale of goods", "total net sales", "net sales",
                         "operating revenue", "total income")),
        ("II. OPERATING EXPENSES", ("cost of revenue", "cost of sales", "cost of materials", "cost of goods",
                         "purchases of stock", "changes in inventories", "fulfillment", "research and development",
                         "technology", "content", "marketing and sales", "sales and marketing", "selling",
                         "general and administrative", "administrative", "employee benefit", "depreciation",
                         "other operating expense", "other expense", "total costs and expenses", "total costs",
                         "total operating expense", "total expense", "operating costs")),
        ("III. OPERATING & NON-OPERATING INCOME", ("income from operations", "loss from operations",
                         "operating income", "operating profit", "interest and other income", "interest income",
                         "finance income", "interest expense", "finance cost", "other income", "other expense, net",
                         "total non-operating", "exceptional")),
        ("IV. TAXES & BOTTOM LINE", ("income before", "before income tax", "before tax", "before provision",
                         "provision for income tax", "provision for tax", "benefit", "tax expense", "current tax",
                         "deferred tax", "income tax", "equity-method", "net income", "net loss",
                         "profit for the year", "profit after tax")),
        ("V.  COMPREHENSIVE INCOME", ("comprehensive",)),
        ("VI. EARNINGS PER SHARE", ("earnings per share", "per share", "weighted-average shares",
                         "weighted-avg shares", "weighted", "basic", "diluted")),
    ],
    "Balance Sheet": [
        ("A.  CURRENT ASSETS", ("cash and cash equivalent", "bank balance", "marketable securities",
                         "current investment", "inventories", "accounts receivable", "trade receivable",
                         "prepaid", "other current asset", "total current assets")),
        ("B.  NON-CURRENT ASSETS", ("property and equipment", "property, plant", "fixed asset", "right-of-use",
                         "operating lease", "goodwill", "intangible", "non-current investment",
                         "deferred tax asset", "other asset", "total non-current assets", "total assets")),
        ("C.  CURRENT LIABILITIES", ("accounts payable", "trade payable", "accrued", "unearned", "short-term debt",
                         "short-term borrowing", "current portion", "other current liabilit", "total current liabilities")),
        ("D.  NON-CURRENT LIABILITIES", ("long-term debt", "long-term borrowing", "long-term lease", "lease liabilit",
                         "deferred tax liabilit", "other long-term", "other non-current liabilit",
                         "total non-current liabilities", "total liabilities")),
        ("E.  STOCKHOLDERS' EQUITY", ("common stock", "preferred stock", "share capital", "additional paid-in",
                         "securities premium", "treasury stock", "retained earnings", "reserves", "accumulated",
                         "other equity", "total stockholders", "total shareholders", "total equity",
                         "total liabilities and", "total equity and liabilities")),
    ],
    "Cash Flow Statement": [
        ("A.  OPERATING ACTIVITIES", ("net income", "profit before tax", "profit for the year", "depreciation",
                         "amortization", "stock-based compensation", "deferred income tax", "non-operating expense",
                         "expense (income), net", "working capital", "changes in", "accounts receivable", "inventories", "accounts payable",
                         "accrued", "unearned", "other assets", "cash generated from operation", "income tax paid",
                         "operating lease assets", "net cash provided by operating", "net cash generated from operating",
                         "net cash from operating", "net cash used in operating", "net cash provided by (used in) operating")),
        ("B.  INVESTING ACTIVITIES", ("purchases of property", "purchase of property", "proceeds from property",
                         "purchases of marketable", "sales and maturities", "acquisitions", "purchase of investment",
                         "net cash used in investing", "net cash provided by (used in) investing",
                         "net cash from investing", "net cash generated from investing")),
        ("C.  FINANCING ACTIVITIES", ("repurchased", "buyback", "proceeds from long-term debt",
                         "repayments of long-term debt", "proceeds from short-term", "repayments of short-term",
                         "principal repayments of finance", "principal repayments of financing", "dividend",
                         "proceeds from issuance", "net cash used in financing",
                         "net cash provided by (used in) financing", "net cash from financing")),
        ("D.  NET CHANGE IN CASH", ("foreign currency effect", "net increase", "net decrease",
                         "cash, cash equivalents", "cash and cash equivalents")),
    ],
}


def _classify_section(statement: str, label: str) -> int:
    """Index of the section this line belongs to (longest keyword match wins).
    -1 means unclassified (rendered after the sections, ungrouped)."""
    lower = label.lower()
    best_idx, best_len = -1, 0
    for idx, (_, keywords) in enumerate(_SECTIONS.get(statement, [])):
        for kw in keywords:
            if kw in lower and len(kw) > best_len:
                best_idx, best_len = idx, len(kw)
    return best_idx


_DROP_LABELS = {"marketing", "december", "particulars",
                "issued shares - and", "outstanding shares - and"}


def _drop_row(label: str) -> bool:
    """Known terse footnote fragments that broadened extraction picks up."""
    return label.strip().lower() in _DROP_LABELS


def _is_noise_label(label: str) -> bool:
    # Only true fragments — keep legitimate one-word lines like "Revenue",
    # "Goodwill", "Inventories", "Fulfillment".
    low = label.strip().lower()
    if len(low) < 4:
        return True
    if low.endswith((" and", " - and", " -", " of", " or", ",")):
        return True
    return False


def _is_subtotal(label: str) -> bool:
    low = label.lower().strip()
    if low.startswith("total") or "net cash" in low:
        return True
    return low in {
        "operating income", "operating profit", "gross profit",
        "income (loss) before income taxes", "income before income taxes", "profit before tax",
        "net income", "net income (loss)", "profit for the year",
        "comprehensive income (loss)", "comprehensive income",
    }


def _number_format(label: str) -> str:
    # Per-share figures keep 2 decimals; everything else is a whole-number punch.
    return "#,##0.00;(#,##0.00)" if "per share" in label.lower() else "#,##0;(#,##0)"


def _as_row(ws, row, label, vals, indent=0, total=False, ncols=0):
    fill = _AS_TOTAL if total else (_AS_ALT if row % 2 == 0 else _AS_WHITE)
    font = _AS_SEC_FONT if total else _AS_NORM_FONT
    border = Border(top=_AS_THIN, bottom=_AS_MED) if total else Border(bottom=_AS_THIN)
    numfmt = _number_format(label)
    c1 = ws.cell(row, 1, "    " * indent + label)
    c1.fill, c1.font, c1.border = fill, font, border
    c1.alignment = Alignment(horizontal="left", vertical="center")
    for i, val in enumerate(vals, 2):
        c = ws.cell(row, i, val)
        c.fill, c.font, c.border = fill, font, border
        c.number_format = numfmt
        c.alignment = Alignment(horizontal="right", vertical="center")


def write_statement_sheet(ws, company: str, filings: list[ParsedPdf], statement: str, periods: list[str]) -> None:
    ws.sheet_view.showGridLines = False
    ncols = len(periods)
    last_col = get_column_letter(ncols + 1)
    ws.column_dimensions["A"].width = 56
    for col in range(2, ncols + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Title block
    ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 26
    c = ws.cell(1, 1, company.upper())
    c.fill, c.font = _AS_HEAD, _AS_TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, ncols + 2):
        ws.cell(1, col).fill = _AS_HEAD
    ws.merge_cells(f"A2:{last_col}2")
    c = ws.cell(2, 1, f"CONSOLIDATED {statement.upper()}")
    c.fill, c.font = _AS_SUBHEAD, _AS_SUB_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    span = f"{periods[0]} to {periods[-1]}" if periods else ""
    ws.merge_cells(f"A3:{last_col}3")
    c = ws.cell(3, 1, f"Multi-Year {span}  |  Figures in millions as reported  |  {company}")
    c.fill, c.font = _AS_NOTE, _AS_NOTE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers
    ws.row_dimensions[5].height = 26
    ws.freeze_panes = "B6"
    h = ws.cell(5, 1, "Particulars")
    h.fill, h.font = _AS_SUBHEAD, _AS_SUB_FONT
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border = Border(top=_AS_MED, bottom=_AS_MED)
    for i, period in enumerate(periods, 2):
        c = ws.cell(5, i, f"{period}\n(Mn)")
        c.fill, c.font = _AS_YEAR, _AS_YEAR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=_AS_MED, bottom=_AS_MED)

    keys = _ordered_keys(filings, statement)
    # group keys by section, preserving statement order within each
    sections = _SECTIONS.get(statement, [])
    buckets: dict[int, list[str]] = {i: [] for i in range(len(sections))}
    last_section = 0
    for key in keys:
        label = _display_label(filings, statement, key)
        if _drop_row(label):
            continue
        idx = _classify_section(statement, label)
        if idx < 0:
            if _is_noise_label(label):
                continue  # true footnote / wrapped-line fragment
            idx = last_section  # inherit the running section, preserving PDF order
        else:
            last_section = idx
        buckets[idx].append(key)
    leftovers: list[str] = []

    row = 6
    for sec_idx, (sec_label, _) in enumerate(sections):
        sec_keys = buckets[sec_idx]
        if not sec_keys:
            continue
        ws.row_dimensions[row].height = 16
        for col in range(1, ncols + 2):
            cell = ws.cell(row, col)
            cell.fill = _AS_SECTION
            if col == 1:
                cell.value = sec_label
                cell.font = _AS_SEC_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        for key in sec_keys:
            label = _display_label(filings, statement, key)
            vals = [_value_for_period(filings, statement, key, p) for p in periods]
            total = _is_subtotal(label)
            _as_row(ws, row, label, vals, indent=0 if total else 1, total=total, ncols=ncols)
            row += 1
    for key in leftovers:
        label = _display_label(filings, statement, key)
        vals = [_value_for_period(filings, statement, key, p) for p in periods]
        _as_row(ws, row, label, vals, indent=1, total=_is_subtotal(label), ncols=ncols)
        row += 1


def write_master_sheet(ws, grouped: dict[str, list[ParsedPdf]], statement: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 56
    companies = sorted(grouped)

    # company -> its period columns; build the flat column plan (with a blank
    # spacer column between companies).
    plan: list[tuple[str, str, int]] = []  # (company, period, col)
    col = 2
    company_spans: list[tuple[str, int, int]] = []
    for company in companies:
        periods = _combined_periods(grouped[company])
        if not periods:
            continue
        start = col
        for period in periods:
            plan.append((company, period, col))
            ws.column_dimensions[get_column_letter(col)].width = 15
            col += 1
        company_spans.append((company, start, col - 1))
        col += 1  # spacer
    last_col_idx = max(col - 2, 2)
    last_col = get_column_letter(last_col_idx)

    # Title block
    ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 26
    c = ws.cell(1, 1, f"MASTER — CONSOLIDATED {statement.upper()}")
    c.fill, c.font = _AS_HEAD, _AS_TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(2, last_col_idx + 1):
        ws.cell(1, cc).fill = _AS_HEAD

    # Company group headers (row 3) + period headers (row 4)
    hc = ws.cell(3, 1, "Particulars")
    hc.fill, hc.font = _AS_SUBHEAD, _AS_SUB_FONT
    hc.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(4, 1).fill = _AS_SUBHEAD
    for company, start, end in company_spans:
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        cc = ws.cell(3, start, company)
        cc.fill, cc.font = _AS_SUBHEAD, _AS_SUB_FONT
        cc.alignment = Alignment(horizontal="center", vertical="center")
    for company, period, ccol in plan:
        c = ws.cell(4, ccol, f"{period}\n(Mn)")
        c.fill, c.font = _AS_YEAR, _AS_YEAR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 26
    ws.freeze_panes = "B5"

    # ordered, classified keys (union across companies)
    keys: list[str] = []
    seen: set[str] = set()
    for company in companies:
        for key in _ordered_keys(grouped[company], statement):
            if key not in seen:
                seen.add(key)
                keys.append(key)

    def disp(key: str) -> str:
        return next((lbl for company in companies
                     if (lbl := _display_label(grouped[company], statement, key)) != key), key)

    sections = _SECTIONS.get(statement, [])
    buckets: dict[int, list[str]] = {i: [] for i in range(len(sections))}
    last_section = 0
    for key in keys:
        label = disp(key)
        if _drop_row(label):
            continue
        idx = _classify_section(statement, label)
        if idx < 0:
            if _is_noise_label(label):
                continue
            idx = last_section
        else:
            last_section = idx
        buckets[idx].append(key)
    leftovers: list[str] = []

    def render(row: int, key: str) -> None:
        label = disp(key)
        total = _is_subtotal(label)
        fill = _AS_TOTAL if total else (_AS_ALT if row % 2 == 0 else _AS_WHITE)
        font = _AS_SEC_FONT if total else _AS_NORM_FONT
        border = Border(top=_AS_THIN, bottom=_AS_MED) if total else Border(bottom=_AS_THIN)
        numfmt = _number_format(label)
        c1 = ws.cell(row, 1, ("" if total else "    ") + label)
        c1.fill, c1.font, c1.border = fill, font, border
        c1.alignment = Alignment(horizontal="left", vertical="center")
        for company, period, ccol in plan:
            c = ws.cell(row, ccol, _value_for_period(grouped[company], statement, key, period))
            c.fill, c.font, c.border = fill, font, border
            c.number_format = numfmt
            c.alignment = Alignment(horizontal="right", vertical="center")

    row = 5
    for sec_idx, (sec_label, _) in enumerate(sections):
        if not buckets[sec_idx]:
            continue
        for cc in range(1, last_col_idx + 1):
            cell = ws.cell(row, cc)
            cell.fill = _AS_SECTION
            if cc == 1:
                cell.value = sec_label
                cell.font = _AS_SEC_FONT
        row += 1
        for key in buckets[sec_idx]:
            render(row, key)
            row += 1
    for key in leftovers:
        render(row, key)
        row += 1


def style_row(ws, row_idx: int, max_col: int, palette: "WorkbookPalette", total: bool = False) -> None:
    fill = palette.total if total else (palette.alt if row_idx % 2 == 0 else palette.white)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.fill = fill
        cell.border = palette.thin_border
        if total:
            cell.font = Font(bold=True, color="1F3864")
        if col_idx > 1:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")


class WorkbookPalette:
    def __init__(self) -> None:
        self.dark = PatternFill("solid", fgColor="1F3864")
        self.mid = PatternFill("solid", fgColor="2E75B6")
        self.alt = PatternFill("solid", fgColor="F2F7FC")
        self.white = PatternFill("solid", fgColor="FFFFFF")
        self.total = PatternFill("solid", fgColor="BDD7EE")
        thin = Side(style="thin", color="B8CCE4")
        medium = Side(style="medium", color="2E75B6")
        self.thin_border = Border(bottom=thin)
        self.header_border = Border(top=medium, bottom=medium)


def build_summary(company: str, filings: list[ParsedPdf]) -> dict[str, dict[str, float | None]]:
    periods = _combined_periods(filings)
    summary = {metric: {} for metric in SUMMARY_LABELS}
    for metric, needles in SUMMARY_LABELS.items():
        for period in periods:
            summary[metric][period] = find_summary_value(filings, metric, needles, period)
    return summary


def find_summary_value(
    filings: list[ParsedPdf], metric: str, needles: tuple[str, ...], period: str
) -> float | None:
    if metric == "Revenue":
        direct = _find_labeled_value(filings, ("Income Statement",), needles, period)
        if direct is not None:
            return direct
        sales = _find_labeled_value(filings, ("Income Statement",), ("sale of goods", "net product sales"), period)
        other_revenue = _find_labeled_value(
            filings,
            ("Income Statement",),
            ("other operating revenues", "net service sales"),
            period,
        )
        if sales is not None and other_revenue is not None:
            return sales + other_revenue
        return sales

    statement_names = {
        "Revenue": ("Income Statement",),
        "PBT": ("Income Statement",),
        "PAT": ("Income Statement",),
        "Total Assets": ("Balance Sheet",),
        "Total Equity": ("Balance Sheet",),
        "Operating Cash Flow": ("Cash Flow Statement",),
    }[metric]
    return _find_labeled_value(filings, statement_names, needles, period)


def _find_labeled_value(
    filings: list[ParsedPdf], statement_names: tuple[str, ...], needles: tuple[str, ...], period: str
) -> float | None:
    for filing in reversed(filings):
        for statement in statement_names:
            for label, values in filing.statements.get(statement, {}).items():
                label_lower = label.lower()
                if period in values and any(needle in label_lower for needle in needles):
                    return values[period]
    return None


def _combined_periods(filings: list[ParsedPdf]) -> list[str]:
    periods: list[str] = []
    for filing in filings:
        periods.extend(filing.periods)
        for statement_rows in filing.statements.values():
            for values in statement_rows.values():
                periods.extend(values.keys())
    return sorted(_dedupe_periods(periods), key=_period_value)


def _dedupe_periods(periods: list[str]) -> list[str]:
    clean = []
    for period in periods:
        if not period:
            continue
        match = re.search(r"20\d{2}", period)
        # Prefer a clean FYxxxx label when a year is present, but keep generic
        # labels (e.g. "Period 1") so values still render when year detection
        # couldn't identify the column headers.
        clean.append(f"FY{match.group(0)}" if match else period)
    return _dedupe(clean)


def _infer_periods_from_rows(statements: dict[str, dict[str, dict[str, float]]]) -> list[str]:
    periods: list[str] = []
    for rows in statements.values():
        for values in rows.values():
            periods.extend(values.keys())
    return _dedupe_periods(periods)


def _period_value(period: str) -> int:
    match = re.search(r"20\d{2}", period)
    return int(match.group(0)) if match else 0


def _period_sort_key(periods: list[str]) -> int:
    values = [_period_value(period) for period in periods]
    return max(values) if values else 0


def _value_for_period(
    filings: list[ParsedPdf], statement: str, label: str, period: str
) -> float | None:
    for filing in reversed(filings):
        value = filing.statements.get(statement, {}).get(label, {}).get(period)
        if value is not None:
            return value
    return None


# Ordered templates that lay each statement out top-to-bottom the way it reads
# on the page. A label is ordered by the LONGEST template phrase it contains, so
# overlaps resolve correctly (e.g. "non-operating income" -> "total non-operating",
# not "operating income").
_STATEMENT_ORDER = {
    "Income Statement": [
        "revenue from operations", "sale of goods", "net product sales",
        "net service sales", "other operating revenue",
        "total net sales", "total revenue", "total income",
        "cost of sales", "cost of materials", "cost of goods", "purchases of stock-in-trade",
        "changes in inventories", "employee benefit", "fulfillment", "technology",
        "sales and marketing", "general and administrative", "depreciation and amorti",
        "other operating expense", "other expense",
        "total operating expense", "total expense",
        "operating income", "operating profit",
        "interest income", "finance income",
        "interest expense", "finance cost",
        "other income", "total non-operating",
        "before exceptional", "exceptional",
        "profit before tax", "before income tax",
        "provision for income tax", "tax expense", "current tax", "deferred tax", "income tax",
        "equity-method", "profit after tax",
        "profit for the year", "net income",
        "other comprehensive", "total comprehensive",
        "basic earnings", "diluted earnings", "earnings per share",
    ],
    "Balance Sheet": [
        "cash and cash equivalent", "bank balance", "marketable securities", "current investment",
        "inventories", "accounts receivable", "trade receivable", "other current asset",
        "total current assets",
        "property and equipment", "property, plant", "fixed asset", "right-of-use", "operating lease",
        "goodwill", "intangible", "non-current investment", "deferred tax asset", "other asset",
        "total non-current assets",
        "total assets",
        "accounts payable", "trade payable", "accrued", "short-term debt", "short-term borrowing",
        "current portion", "other current liabilit", "total current liabilities",
        "long-term debt", "long-term borrowing", "long-term lease", "lease liabilit",
        "deferred tax liabilit", "other long-term", "other non-current liabilit",
        "total non-current liabilities", "total liabilities",
        "common stock", "share capital", "additional paid-in", "securities premium", "treasury stock",
        "retained earnings", "reserves", "accumulated", "other equity",
        "total stockholders", "total shareholders", "total equity", "total equity and liabilities",
    ],
    "Cash Flow Statement": [
        "profit before tax", "net income", "profit for the year",
        "depreciation", "amortization", "stock-based compensation", "stock based compensation",
        "changes in", "accounts receivable", "inventories", "accounts payable",
        "cash generated from operation", "income tax paid",
        "net cash provided by operating", "net cash generated from operating",
        "net cash from operating", "net cash used in operating",
        "purchases of property", "purchase of property", "purchases of marketable", "sales and maturities",
        "acquisitions", "purchase of investment",
        "net cash used in investing", "net cash provided by (used in) investing",
        "net cash from investing", "net cash generated from investing",
        "proceeds from long-term debt", "repayments of long-term debt",
        "proceeds from short-term", "repayments of short-term",
        "principal repayments of finance", "principal repayments of financing", "dividend", "buyback",
        "net cash used in financing", "net cash provided by (used in) financing", "net cash from financing",
        "foreign currency effect", "net increase", "net decrease",
        "cash, cash equivalents", "cash and cash equivalents",
    ],
}


def _ordered_keys(filings: list[ParsedPdf], statement: str) -> list[str]:
    """Canonical keys ordered by each line's AVERAGE normalized position across
    the filings it appears in. Using the average (rather than a single filing's
    order) keeps the PDF's row order even when a given year's report didn't parse
    every line, so nothing gets dumped at the bottom."""
    positions: dict[str, list[float]] = {}
    for filing in filings:
        order = filing.row_order.get(statement, {})
        if not order:
            continue
        span = max(max(order.values()), 1)
        for key, idx in order.items():
            positions.setdefault(key, []).append(idx / span)
    return sorted(positions, key=lambda k: sum(positions[k]) / len(positions[k]))


def _display_label(filings: list[ParsedPdf], statement: str, key: str) -> str:
    """The verbatim label as written on the PDF, preferring the newest filing."""
    for filing in reversed(filings):
        label = filing.display_labels.get(statement, {}).get(key)
        if label:
            return label
    return key


def _preferred_label_order(statement: str, label: str) -> tuple[int, str]:
    template = _STATEMENT_ORDER.get(statement, [])
    lower = label.lower()
    best_idx, best_len = None, -1
    for idx, phrase in enumerate(template):
        if phrase in lower and len(phrase) > best_len:
            best_idx, best_len = idx, len(phrase)
    if best_idx is None:
        return (len(template) + 1, lower)
    return (best_idx, lower)


def _is_total_label(label: str) -> bool:
    lower = label.lower()
    return lower.startswith("total") or "net cash" in lower or lower in {
        "profit for the year",
        "net income",
        "profit before tax",
        "income before income taxes",
    }


def safe_filename(name: str) -> str:
    name = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
    return name or f"Company_{uuid.uuid4().hex[:8]}"


def _dedupe(items: Iterable[str]) -> list[str]:
    seen = set()
    out = []
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def create_processing_dir(base: Path) -> Path:
    path = base / uuid.uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    return path


def cleanup_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def temporary_workspace() -> Path:
    return Path(tempfile.mkdtemp(prefix="pdf_financials_"))
